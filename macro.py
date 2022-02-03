import sys
import datetime

import numpy as np
import openpyxl
import pandas as pd
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QGridLayout, QFileDialog, QLabel
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.descriptors import (
    String,
    Sequence,
    Integer,
)
from openpyxl.descriptors.serialisable import Serialisable

BUILTIN_FORMATS = {
    0: 'General',
    1: '0',
    2: '0.00',
    3: '#,##0',
    4: '#,##0.00',
    5: '"$"#,##0_);("$"#,##0)',
    6: '"$"#,##0_);[Red]("$"#,##0)',
    7: '"$"#,##0.00_);("$"#,##0.00)',
    8: '"$"#,##0.00_);[Red]("$"#,##0.00)',
    9: '0%',
    10: '0.00%',
    11: '0.00E+00',
    12: '# ?/?',
    13: '# ??/??',
    14: 'mm-dd-yy',
    15: 'd-mmm-yy',
    16: 'd-mmm',
    17: 'mmm-yy',
    18: 'h:mm AM/PM',
    19: 'h:mm:ss AM/PM',
    20: 'h:mm',
    21: 'h:mm:ss',
    22: 'm/d/yy h:mm',

    37: '#,##0_);(#,##0)',
    38: '#,##0_);[Red](#,##0)',
    39: '#,##0.00_);(#,##0.00)',
    40: '#,##0.00_);[Red](#,##0.00)',

    41: r'_(* #,##0_);_(* \(#,##0\);_(* "-"_);_(@_)',
    42: r'_("$"* #,##0_);_("$"* \(#,##0\);_("$"* "-"_);_(@_)',
    43: r'_(* #,##0.00_);_(* \(#,##0.00\);_(* "-"??_);_(@_)',

    44: r'_("$"* #,##0.00_)_("$"* \(#,##0.00\)_("$"* "-"??_)_(@_)',
    45: 'mm:ss',
    46: '[h]:mm:ss',
    47: 'mmss.0',
    48: '##0.0E+0',
    49: '@', }

BUILTIN_FORMATS_MAX_SIZE = 164
BUILTIN_FORMATS_REVERSE = dict(
        [(value, key) for key, value in BUILTIN_FORMATS.items()])


#현재시간 파일 저장 o
#사기성총액 - 지불소계 o
#사기성총액 / 월별 지급액 / 지불소계 총합 / 대비 총합
#지불집계표 : 비고(업체정보) / 매입 : 품목비교 삭제 o
#매입 Sheet : 작성일자 오르차순 정리
#지불집계표 Sheet : 상호명으로 오르차순 정리
#Columns 크기(가로만) 조정
#표경계선 그리기


class QtGUI(QWidget):

    def __init__(self):
        super().__init__()
        self.setWindowTitle("매크로")
        self.resize(200, 200)
        self.qclist = []
        self.position = 0
        self.Lgrid = QGridLayout()
        self.setLayout(self.Lgrid)
        self.label1 = QLabel('', self)
        # self.label2 = QLabel('', self)
        addbutton1 = QPushButton('Open File', self)
        self.Lgrid.addWidget(self.label1, 1, 1)
        self.Lgrid.addWidget(addbutton1, 2, 1)
        addbutton1.clicked.connect(self.add_open)
        # addbutton2 = QPushButton('Divide Sheet', self)
        # self.Lgrid.addWidget(self.label2, 3, 1)
        # self.Lgrid.addWidget(addbutton2, 4, 1)
        # addbutton2.clicked.connect(self.div_sheet)
        self.show()

    def autoFitColumnSize(self, worksheet, columns=None, margin=2):
        for i, column_cells in enumerate(worksheet.columns):
            is_ok = False
            if columns is None:
                is_ok = True
            elif isinstance(columns, list) and i in columns:
                is_ok = True
            if is_ok:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + margin
        return worksheet

    def add_open(self):
        self.FileOpen = QFileDialog.getOpenFileName(self, 'Open file', './')
        self.label1.setText(self.FileOpen[0])
        self.wb = openpyxl.load_workbook(self.FileOpen[0])
        self.rowDf = pd.read_excel(io=self.FileOpen[0])
        self.rowDf['실결제금액'] = self.rowDf['실결제금액'].fillna(0)
        self.rowDf = self.rowDf.astype({'상호': 'string', '대표자명': 'string',
                                        '주소': 'string', '합계금액': 'int64',
                                        '공급가액': 'int64', '세액': 'int64',
                                        '품목명': 'string', '비고(업체정보)': 'string',
                                        '실결제금액': 'int64', '결제수단': 'string',
                                        '차액': 'int64', '현장': 'string'
                                        })
        self.rowDf['작성일자(년)'] = self.rowDf['작성일자'].dt.year
        self.rowDf['작성일자(월)'] = self.rowDf['작성일자'].dt.month
        self.groupbydf = self.rowDf.groupby(['작성일자(년)', '작성일자(월)', '공급자사업자등록번호', '상호', '대표자명']).agg('sum')
        self.df_static = pd.DataFrame(self.groupbydf)
        self.df_static = self.df_static.reset_index()
        years = set(self.df_static['작성일자(년)'].values)
        for year in years:
            static = []
            year_df = self.df_static[self.df_static['작성일자(년)'] == year]
            enterprises_nums = set(year_df['공급자사업자등록번호'].values)
            for enterprise_num in enterprises_nums:
                # 사업자번호
                enterprise_num_df = year_df[year_df['공급자사업자등록번호'] == enterprise_num]
                dataset = np.zeros(shape=(19,), dtype=object)
                index = enterprise_num_df['공급자사업자등록번호'].index[0]
                np.put(dataset, [0], enterprise_num_df['공급자사업자등록번호'][index])
                np.put(dataset, [1], enterprise_num_df['상호'][index])
                np.put(dataset, [2], enterprise_num_df['대표자명'][index])
                real_total = 0
                for idx, row in enumerate(enterprise_num_df.iterrows()):
                    row = row[1]
                    real_total += row['실결제금액']
                    if idx == 0:
                        np.put(dataset, [row['작성일자(월)'] + 3], row['합계금액'])
                        np.put(dataset, [16], row['실결제금액'])
                    else:
                        np.put(dataset, [row['작성일자(월)'] + 3], row['합계금액'])
                        np.put(dataset, [16], real_total)
                static.append(dataset)

            self.df_result = pd.DataFrame(static,
                                          columns=['사업자번호', '상호', '대표자', '사기성총액', '1월', '2월', '3월', '4월', '5월', '6월',
                                                   '7월', '8월', '9월', '10월', '11월', '12월', '지불소계', '대비', '비고(업체정보)'])
            self.df_result['사기성총액'] = self.df_result[
                ['1월', '2월', '3월', '4월', '5월', '6월', '7월', '8월', '9월', '10월', '11월', '12월']].sum(axis=1)
            self.df_result['대비'] = self.df_result['사기성총액'] - self.df_result['지불소계']
            self.df_result['비고(업체정보)'] = self.df_result[(self.df_result['비고(업체정보)'] == 0)] = '-'

            ws = self.wb.create_sheet(str(year) + '_지불집계표')
            for r in dataframe_to_rows(self.df_result, index=True, header=True):
                print(r)
                ws.append(r)
            # ws.delete_rows(2)
            for num in range(1, 30):
                ws.cell(1, num).font = Font(bold=True)

            ws['B2'].value = "합계"
            ws['T2'].value = "-"
            ws.merge_cells("B2:D2")
            for num in range(1, 30):
                ws.cell(2, num).font = Font(bold=True)

            ws['E2'] = '=SUM(E3:E1000)'
            ws['F2'] = '=SUM(F3:F1000)'
            ws['G2'] = '=SUM(G3:G1000)'
            ws['H2'] = '=SUM(H3:H1000)'
            ws['I2'] = '=SUM(I3:I1000)'
            ws['J2'] = '=SUM(J3:J1000)'
            ws['L2'] = '=SUM(L3:L1000)'
            ws['M2'] = '=SUM(M3:M1000)'
            ws['N2'] = '=SUM(N3:N1000)'
            ws['O2'] = '=SUM(O3:O1000)'
            ws['P2'] = '=SUM(P3:P1000)'
            ws['Q2'] = '=SUM(Q3:Q1000)'
            ws['R2'] = '=SUM(R3:R1000)'
            ws['S2'] = '=SUM(S3:S1000)'
            # 너비
            ws.column_dimensions['A'].width = 4
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 19
            ws.column_dimensions['D'].width = 13
            ws.column_dimensions['E'].width = 13
            ws.column_dimensions['F'].width = 13
            ws.column_dimensions['G'].width = 13
            ws.column_dimensions['H'].width = 13
            ws.column_dimensions['I'].width = 13
            ws.column_dimensions['J'].width = 13
            ws.column_dimensions['K'].width = 13
            ws.column_dimensions['L'].width = 13
            ws.column_dimensions['M'].width = 13
            ws.column_dimensions['N'].width = 13
            ws.column_dimensions['O'].width = 13
            ws.column_dimensions['P'].width = 13
            ws.column_dimensions['Q'].width = 13
            ws.column_dimensions['R'].width = 13
            ws.column_dimensions['S'].width = 13
            ws.column_dimensions['T'].width = 14

            # 데이터 서식
            for rng in ws['E2':'S100']:
                for cell in rng:
                    cell.number_format = '#,##0'

            # 가운데 정렬


            # 표 선그리기


        self.wb.save(self.FileOpen[0].split('.')[0] + "_집계_"+datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S')+".xlsx")

    # def div_sheet(self):
    #     self.FileOpen = QFileDialog.getOpenFileName(self, 'Divide Sheet', './')
    #     self.label2.setText(self.FileOpen[0])
    #     self.wb = openpyxl.load_workbook(self.FileOpen[0])
    #     self.rowDf_divide = pd.read_excel(io=self.FileOpen[0])
    #     self.rowDf_divide['실결제금액'] = self.rowDf_divide['실결제금액'].fillna(0)
    #     self.rowDf_divide = self.rowDf_divide.astype({'상호': 'string', '대표자명': 'string',
    #                                                   '주소': 'string', '합계금액': 'int64',
    #                                                   '공급가액': 'int64', '세액': 'int64',
    #                                                   '품목명': 'string', '품목비고': 'string',
    #                                                   '실결제금액': 'int64', '결제수단': 'string',
    #                                                   '차액': 'int64', '현장': 'string'
    #                                                   })
    #     print(self.rowDf_divide)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = QtGUI()
    app.exec_()
